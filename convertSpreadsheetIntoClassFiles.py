from openpyxl import load_workbook
import pandas as pd
import os

filename = 'GoogleOpenImages-allClasses.xlsx'

wb = load_workbook(filename, data_only=True)
print(wb)

datasheets = []

for sheetName in wb.sheetnames:
    if 'draft' not in sheetName:
        datasheets.append(sheetName)

        if 'OID' in sheetName:
            datasheets.append('OID-extra')

modelLists = dict()

for sheet in datasheets:
    if 'OID-extra' in sheet:
        # Read OID
        df = pd.read_excel(filename, sheet_name='OID')
    else:
        df = pd.read_excel(filename, sheet_name=sheet)

    cols = df.columns

    models = []
    appendModels = False

    for col in cols:
        if appendModels:
            models.append(col)

        if 'rename' in col.lower():
            # The models are columns after the "rename" column
            appendModels = True        


    print(df.to_string())

    for model in models:
        if 'OID-extra' in sheet:
            n = pd.to_numeric(df[model])            
            subset = df[n > 1]
        else:
            subset = df[df[model].notna()]


        # Only include columns to "Rename" column
        reducedCols = subset.loc[:, :'Rename']

        # Insert 'database' column
        reducedCols['database'] = sheet

        if model not in modelLists:
            # First time that we encounter this model
            modelLists[model] = reducedCols
        else:
            modelLists[model] = pd.concat([modelLists[model], reducedCols], 
                                          axis=0, 
                                          ignore_index=True)

        print(modelLists[model].to_string())

os.makedirs('datasets', exist_ok=True)

for model, df in modelLists.items():
    df.to_csv(os.path.join('datasets', '{}.csv'.format(model)), na_rep='')
    
    



